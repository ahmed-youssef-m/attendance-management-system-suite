from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def append_data_to_excel(file_path, data):
    try:
        # Check if the file exists
        file_exists = True
        try:
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            file_exists = False

        if not file_exists:
            # Create a new workbook if the file doesn't exist
            workbook = Workbook()
            # Create a default sheet
            sheet = workbook.active
            sheet.title = 'Data'
            # Add headers (optional)
            headers = ['Student ID']
            sheet.append(headers)
            # Apply bold font to headers
            for cell in sheet[1]:
                cell.font = Font(bold=True,size=10)

        # Select the active sheet
        sheet = workbook.active

        # Determine the next available row for appending data
        next_row = sheet.max_row + 1

        # Write the data row by row
        for row_data in data:
            sheet.append(row_data)

        # Save the workbook to the specified file path
        workbook.save(file_path)

        print("Data appended successfully.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


# # Example usage
# excel_file_path = 'attend.xlsx'
#
# # Data to append (list of rows, where each row is a list of values)
# data_to_append = [
#     [0]
# ]
#
# # Append data to the Excel file, create it if it doesn't exist
# append_data_to_excel(excel_file_path, data_to_append)
